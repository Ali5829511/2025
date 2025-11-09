"""
Main Flask server for Faculty Housing Management System
خادم Flask الرئيسي لنظام إدارة إسكان أعضاء هيئة التدريس
"""

from flask import Flask, request, jsonify, send_from_directory, make_response, abort, send_file
from flask_cors import CORS
from werkzeug.security import safe_join
from werkzeug.utils import secure_filename
import os
from dotenv import load_dotenv
import database
import auth
import plate_recognizer
import parkpow_integration
import car_image_analyzer
import car_data_exporter
import vehicle_report_exporter
import import_historical_vehicles
import housing_report_generator
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

# Load environment variables from .env file
load_dotenv()

# Initialize Flask app
app = Flask(__name__)

# CORS Configuration for production deployment
# إعدادات CORS للنشر على الإنتاج

# Get deployment URL from environment or use default
DEPLOYMENT_URL = os.environ.get('DEPLOYMENT_URL', 'https://housing-management-system-83yt.onrender.com')

# Default allowed origins
ALLOWED_ORIGINS = [
    DEPLOYMENT_URL,
    'http://localhost:5000',
    'http://127.0.0.1:5000'
]

# Validate and add additional origins from environment variable if specified
def validate_origin(origin):
    """Validate that an origin is a properly formatted URL"""
    try:
        from urllib.parse import urlparse
        parsed = urlparse(origin)
        # Must have scheme (http/https) and netloc (domain)
        return parsed.scheme in ('http', 'https') and bool(parsed.netloc)
    except Exception:
        return False

env_origins = os.environ.get('ALLOWED_ORIGINS', '')
if env_origins:
    for origin in env_origins.split(','):
        origin = origin.strip()
        if origin and validate_origin(origin):
            ALLOWED_ORIGINS.append(origin)
        elif origin:
            app.logger.warning(f'Ignored invalid origin from ALLOWED_ORIGINS: {origin}')

# Configure CORS with proper settings for production
CORS(app, 
     supports_credentials=True,
     origins=ALLOWED_ORIGINS,
     allow_headers=['Content-Type', 'Authorization'],
     methods=['GET', 'POST', 'PUT', 'DELETE', 'OPTIONS'])

# Configuration
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('FLASK_ENV') == 'production'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size

# Base directory for serving files
BASE_DIR = os.path.abspath('.')

# Upload configuration
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads', 'car_images')
THUMBNAIL_FOLDER = os.path.join(BASE_DIR, 'uploads', 'thumbnails')
EXPORT_FOLDER = os.path.join(BASE_DIR, 'exports')

# Create directories if they don't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(THUMBNAIL_FOLDER, exist_ok=True)
os.makedirs(EXPORT_FOLDER, exist_ok=True)

# Allowed file extensions
ALLOWED_EXTENSIONS = {'.html', '.css', '.js', '.png', '.jpg', '.jpeg', '.gif', '.svg', '.pdf', '.md'}
ALLOWED_IMAGE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif'}

def is_safe_path(filename):
    """Check if the requested file is safe to serve"""
    if '..' in filename or filename.startswith('/'):
        return False
    
    blocked_files = {'.env', '.git', '__pycache__', 'requirements.txt', '.gitignore', '.py', '.db'}
    if any(blocked in filename for blocked in blocked_files):
        return False
    
    _, ext = os.path.splitext(filename)
    if ext.lower() not in ALLOWED_EXTENSIONS:
        return False
    
    return True


def allowed_image_file(filename):
    """Check if file is an allowed image type"""
    _, ext = os.path.splitext(filename)
    return ext.lower() in ALLOWED_IMAGE_EXTENSIONS

# Initialize database on startup
with app.app_context():
    database.init_database()

# ==================== Authentication Routes ====================

@app.route('/api/auth/login', methods=['POST'])
def login():
    """Login endpoint"""
    try:
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        
        if not username or not password:
            return jsonify({
                'success': False,
                'error': 'Username and password required',
                'error_ar': 'اسم المستخدم وكلمة المرور مطلوبان'
            }), 400
        
        # Verify credentials
        user = database.verify_user(username, password)
        
        if not user:
            # Log failed login attempt
            database.log_audit(
                None,
                f'Failed login attempt for username: {username}',
                ip_address=request.remote_addr
            )
            return jsonify({
                'success': False,
                'error': 'Invalid credentials',
                'error_ar': 'اسم المستخدم أو كلمة المرور غير صحيحة'
            }), 401
        
        # Create session
        session_token, expires_at = auth.create_session(
            user['id'],
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent')
        )
        
        # Log successful login
        database.log_audit(
            user['id'],
            'User logged in',
            ip_address=request.remote_addr
        )
        
        # Create response
        response = make_response(jsonify({
            'success': True,
            'session_token': session_token,
            'user': {
                'id': user['id'],
                'username': user['username'],
                'name': user['name'],
                'role': user['role'],
                'email': user['email']
            },
            'message': 'Login successful',
            'message_ar': 'تم تسجيل الدخول بنجاح'
        }))
        
        # Set session cookie
        response.set_cookie(
            'session_token',
            session_token,
            expires=expires_at,
            httponly=True,
            secure=app.config['SESSION_COOKIE_SECURE'],
            samesite=app.config['SESSION_COOKIE_SAMESITE']
        )
        
        return response
        
    except Exception as e:
        app.logger.error(f'Login error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Internal server error',
            'error_ar': 'خطأ في الخادم'
        }), 500

@app.route('/api/auth/logout', methods=['POST'])
@auth.require_auth
def logout():
    """Logout endpoint"""
    try:
        session_token = request.cookies.get('session_token')
        
        if session_token:
            auth.destroy_session(session_token)
            
            # Log logout
            database.log_audit(
                request.user['id'],
                'User logged out',
                ip_address=request.remote_addr
            )
        
        response = make_response(jsonify({
            'success': True,
            'message': 'Logged out successfully',
            'message_ar': 'تم تسجيل الخروج بنجاح'
        }))
        
        # Clear session cookie
        response.set_cookie('session_token', '', expires=0)
        
        return response
        
    except Exception as e:
        app.logger.error(f'Logout error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Internal server error',
            'error_ar': 'خطأ في الخادم'
        }), 500

@app.route('/api/auth/validate', methods=['GET'])
def validate():
    """Validate session endpoint"""
    try:
        session_token = request.cookies.get('session_token')
        
        if not session_token:
            return jsonify({
                'success': False,
                'authenticated': False
            }), 401
        
        user = auth.validate_session(session_token)
        
        if not user:
            return jsonify({
                'success': False,
                'authenticated': False
            }), 401
        
        return jsonify({
            'success': True,
            'authenticated': True,
            'user': {
                'id': user['id'],
                'username': user['username'],
                'name': user['name'],
                'role': user['role'],
                'email': user['email']
            }
        })
        
    except Exception as e:
        app.logger.error(f'Validation error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Internal server error',
            'error_ar': 'خطأ في الخادم'
        }), 500

@app.route('/api/auth/change-password', methods=['POST'])
@auth.require_auth
def change_password():
    """Change password endpoint"""
    try:
        data = request.get_json()
        current_password = data.get('current_password')
        new_password = data.get('new_password')
        
        if not current_password or not new_password:
            return jsonify({
                'success': False,
                'error': 'Current and new password required',
                'error_ar': 'كلمة المرور الحالية والجديدة مطلوبة'
            }), 400
        
        # Verify current password
        user = database.verify_user(request.user['username'], current_password)
        if not user:
            return jsonify({
                'success': False,
                'error': 'Current password is incorrect',
                'error_ar': 'كلمة المرور الحالية غير صحيحة'
            }), 401
        
        # Update password
        success = database.update_user_password(request.user['id'], new_password)
        
        if success:
            # Log password change
            database.log_audit(
                request.user['id'],
                'Password changed',
                ip_address=request.remote_addr
            )
            
            return jsonify({
                'success': True,
                'message': 'Password changed successfully',
                'message_ar': 'تم تغيير كلمة المرور بنجاح'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Failed to change password',
                'error_ar': 'فشل تغيير كلمة المرور'
            }), 500
            
    except Exception as e:
        app.logger.error(f'Change password error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Internal server error',
            'error_ar': 'خطأ في الخادم'
        }), 500

# ==================== Plate Recognition API ====================

@app.route('/api/plate-recognizer/status', methods=['GET'])
@auth.require_auth
def plate_recognizer_status():
    """Get Plate Recognizer API status"""
    try:
        status = plate_recognizer.get_api_status()
        return jsonify(status)
    except Exception as e:
        app.logger.error(f'Plate recognizer status error: {str(e)}')
        return jsonify({
            'success': False,
            'error': str(e),
            'error_ar': 'خطأ في التحقق من حالة الخدمة'
        }), 500

@app.route('/api/plate-recognizer/recognize', methods=['POST'])
@auth.require_auth
def recognize_plate():
    """Recognize license plate from uploaded image"""
    try:
        # Check if image is provided
        if 'image' not in request.files and 'base64_image' not in request.json:
            return jsonify({
                'success': False,
                'error': 'No image provided',
                'error_ar': 'لم يتم تقديم صورة'
            }), 400
        
        regions = request.form.get('regions', 'sa').split(',') if 'image' in request.files else None
        if request.json and 'regions' in request.json:
            regions = request.json['regions']
        
        # Handle file upload
        if 'image' in request.files:
            image_file = request.files['image']
            
            # Read image bytes
            image_bytes = image_file.read()
            
            # Recognize plate
            result = plate_recognizer.recognize_plate_from_bytes(image_bytes, regions)
        
        # Handle base64 image
        elif request.json and 'base64_image' in request.json:
            base64_image = request.json['base64_image']
            result = plate_recognizer.recognize_plate_from_base64(base64_image, regions)
        
        else:
            return jsonify({
                'success': False,
                'error': 'Invalid request',
                'error_ar': 'طلب غير صالح'
            }), 400
        
        # If recognition was successful, log it and check against database
        if result.get('success') and result.get('results'):
            for plate_data in result['results']:
                plate_number = plate_data['plate']
                confidence = plate_data['confidence']
                
                # Find vehicle in database
                vehicle = plate_recognizer.find_vehicle_by_plate(plate_number)
                
                # Log recognition
                vehicle_id = vehicle['id'] if vehicle else None
                plate_recognizer.log_plate_recognition(
                    request.user['id'],
                    plate_number,
                    confidence,
                    vehicle_id
                )
                
                # Add vehicle info to result
                plate_data['vehicle_info'] = vehicle
        
        # Log audit
        database.log_audit(
            request.user['id'],
            'Plate recognition performed',
            ip_address=request.remote_addr
        )
        
        return jsonify(result)
    
    except Exception as e:
        app.logger.error(f'Plate recognition error: {str(e)}')
        return jsonify({
            'success': False,
            'error': str(e),
            'error_ar': 'خطأ في تمييز اللوحة'
        }), 500

@app.route('/api/plate-recognizer/history', methods=['GET'])
@auth.require_auth
def plate_recognition_history():
    """Get plate recognition history"""
    try:
        # Get query parameters
        limit = request.args.get('limit', 50, type=int)
        offset = request.args.get('offset', 0, type=int)
        
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Get recognition history with user and vehicle info
        cursor.execute('''
            SELECT 
                p.*,
                u.username,
                u.name as user_name,
                v.plate_number as registered_plate,
                v.make,
                v.model,
                r.name as owner_name,
                r.unit_number
            FROM plate_recognition_log p
            LEFT JOIN users u ON p.user_id = u.id
            LEFT JOIN vehicles v ON p.vehicle_id = v.id
            LEFT JOIN residents r ON v.owner_id = r.id
            ORDER BY p.recognized_at DESC
            LIMIT ? OFFSET ?
        ''', (limit, offset))
        
        rows = cursor.fetchall()
        
        # Get total count
        cursor.execute('SELECT COUNT(*) FROM plate_recognition_log')
        total = cursor.fetchone()[0]
        
        conn.close()
        
        history = [dict(row) for row in rows]
        
        return jsonify({
            'success': True,
            'history': history,
            'total': total,
            'limit': limit,
            'offset': offset
        })
    
    except Exception as e:
        app.logger.error(f'Plate recognition history error: {str(e)}')
        return jsonify({
            'success': False,
            'error': str(e),
            'error_ar': 'خطأ في جلب سجل التمييز'
        }), 500

@app.route('/api/plate-recognizer/export-excel', methods=['GET'])
@auth.require_auth
def export_plate_recognition_excel():
    """Export plate recognition history to Excel"""
    try:
        # Get query parameters
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Build query with optional date filters
        query = '''
            SELECT 
                p.id,
                p.plate_number,
                p.confidence,
                p.recognized_at,
                u.name as user_name,
                v.plate_number as registered_plate,
                v.vehicle_type,
                v.make,
                v.model,
                v.year,
                v.color,
                v.sticker_number,
                r.name as owner_name,
                r.national_id,
                r.phone as owner_phone,
                r.department,
                r.job_title,
                r.unit_number,
                b.name as building_name
            FROM plate_recognition_log p
            LEFT JOIN users u ON p.user_id = u.id
            LEFT JOIN vehicles v ON p.vehicle_id = v.id
            LEFT JOIN residents r ON v.owner_id = r.id
            LEFT JOIN buildings b ON r.building_id = b.id
        '''
        
        params = []
        if start_date and end_date:
            query += ' WHERE DATE(p.recognized_at) BETWEEN ? AND ?'
            params.extend([start_date, end_date])
        elif start_date:
            query += ' WHERE DATE(p.recognized_at) >= ?'
            params.append(start_date)
        elif end_date:
            query += ' WHERE DATE(p.recognized_at) <= ?'
            params.append(end_date)
        
        query += ' ORDER BY p.recognized_at DESC'
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        conn.close()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير تمييز اللوحات"
        
        # Define styles
        header_fill = PatternFill(start_color="1A5F7A", end_color="1A5F7A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12, name="Arial")
        cell_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        
        # Headers in Arabic
        headers = [
            'م',  # Number
            'رقم اللوحة',  # Plate Number
            'نسبة الدقة %',  # Confidence
            'التاريخ والوقت',  # Date/Time
            'المستخدم',  # User
            'اسم المالك',  # Owner Name
            'الرقم الوطني',  # National ID
            'الهاتف',  # Phone
            'القسم',  # Department
            'المسمى الوظيفي',  # Job Title
            'رقم الوحدة',  # Unit Number
            'اسم المبنى',  # Building Name
            'نوع المركبة',  # Vehicle Type
            'الماركة',  # Make
            'الموديل',  # Model
            'السنة',  # Year
            'اللون',  # Color
            'رقم الملصق',  # Sticker Number
            'حالة التسجيل'  # Registration Status
        ]
        
        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = cell_border
            cell.alignment = center_alignment
        
        # Set column widths
        column_widths = [5, 15, 12, 18, 15, 20, 15, 15, 20, 20, 12, 15, 15, 12, 12, 8, 10, 15, 15]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # Write data
        for idx, row in enumerate(rows, start=2):
            # Row number
            cell = ws.cell(row=idx, column=1)
            cell.value = idx - 1
            cell.alignment = center_alignment
            cell.border = cell_border
            
            # Plate number
            cell = ws.cell(row=idx, column=2)
            cell.value = row['plate_number']
            cell.alignment = center_alignment
            cell.border = cell_border
            cell.font = Font(bold=True, size=11)
            
            # Confidence
            cell = ws.cell(row=idx, column=3)
            confidence_percent = round((row['confidence'] or 0) * 100, 1)
            cell.value = confidence_percent
            cell.alignment = center_alignment
            cell.border = cell_border
            
            # DateTime
            cell = ws.cell(row=idx, column=4)
            if row['recognized_at']:
                try:
                    dt = datetime.fromisoformat(row['recognized_at'])
                    cell.value = dt.strftime('%Y-%m-%d %H:%M:%S')
                except (ValueError, TypeError):
                    # Fallback for different datetime formats
                    cell.value = str(row['recognized_at'])
            cell.alignment = center_alignment
            cell.border = cell_border
            
            # User
            cell = ws.cell(row=idx, column=5)
            cell.value = row['user_name'] or '-'
            cell.alignment = right_alignment
            cell.border = cell_border
            
            # Owner Name
            cell = ws.cell(row=idx, column=6)
            cell.value = row['owner_name'] or 'غير مسجل'
            cell.alignment = right_alignment
            cell.border = cell_border
            if not row['owner_name']:
                cell.font = Font(color="FF0000")
            
            # National ID
            cell = ws.cell(row=idx, column=7)
            cell.value = row['national_id'] or '-'
            cell.alignment = center_alignment
            cell.border = cell_border
            
            # Phone
            cell = ws.cell(row=idx, column=8)
            cell.value = row['owner_phone'] or '-'
            cell.alignment = center_alignment
            cell.border = cell_border
            
            # Department
            cell = ws.cell(row=idx, column=9)
            cell.value = row['department'] or '-'
            cell.alignment = right_alignment
            cell.border = cell_border
            
            # Job Title
            cell = ws.cell(row=idx, column=10)
            cell.value = row['job_title'] or '-'
            cell.alignment = right_alignment
            cell.border = cell_border
            
            # Unit Number
            cell = ws.cell(row=idx, column=11)
            cell.value = row['unit_number'] or '-'
            cell.alignment = center_alignment
            cell.border = cell_border
            
            # Building Name
            cell = ws.cell(row=idx, column=12)
            cell.value = row['building_name'] or '-'
            cell.alignment = right_alignment
            cell.border = cell_border
            
            # Vehicle Type
            cell = ws.cell(row=idx, column=13)
            cell.value = row['vehicle_type'] or '-'
            cell.alignment = right_alignment
            cell.border = cell_border
            
            # Make
            cell = ws.cell(row=idx, column=14)
            cell.value = row['make'] or '-'
            cell.alignment = right_alignment
            cell.border = cell_border
            
            # Model
            cell = ws.cell(row=idx, column=15)
            cell.value = row['model'] or '-'
            cell.alignment = right_alignment
            cell.border = cell_border
            
            # Year
            cell = ws.cell(row=idx, column=16)
            cell.value = row['year'] or '-'
            cell.alignment = center_alignment
            cell.border = cell_border
            
            # Color
            cell = ws.cell(row=idx, column=17)
            cell.value = row['color'] or '-'
            cell.alignment = right_alignment
            cell.border = cell_border
            
            # Sticker Number
            cell = ws.cell(row=idx, column=18)
            cell.value = row['sticker_number'] or '-'
            cell.alignment = center_alignment
            cell.border = cell_border
            
            # Registration Status
            cell = ws.cell(row=idx, column=19)
            status = 'مسجلة' if row['owner_name'] else 'غير مسجلة'
            cell.value = status
            cell.alignment = center_alignment
            cell.border = cell_border
            if row['owner_name']:
                cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                cell.font = Font(color="155724", bold=True)
            else:
                cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                cell.font = Font(color="721C24", bold=True)
        
        # Freeze first row
        ws.freeze_panes = 'A2'
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'تقرير_تمييز_اللوحات_{timestamp}.xlsx'
        
        # Log audit
        database.log_audit(
            request.user['id'],
            f'Exported plate recognition report ({len(rows)} records)',
            ip_address=request.remote_addr
        )
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        # Log error with minimal sensitive information
        app.logger.error(f'Excel export error: {type(e).__name__}', 
                        extra={'error_type': type(e).__name__,
                               'timestamp': datetime.now().isoformat()})
        return jsonify({
            'success': False,
            'error': 'Failed to export report. Please contact system administrator.',
            'error_ar': 'فشل تصدير التقرير. يرجى الاتصال بمسؤول النظام.'
        }), 500

# ==================== ParkPow API Integration ====================

@app.route('/api/parkpow/status', methods=['GET'])
@auth.require_auth
def parkpow_status():
    """Get ParkPow API status"""
    try:
        status = parkpow_integration.get_api_status()
        return jsonify(status)
    except Exception as e:
        app.logger.error(f'ParkPow status error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Failed to check ParkPow status',
            'error_ar': 'خطأ في التحقق من حالة خدمة ParkPow'
        }), 500


@app.route('/api/parkpow/recognize', methods=['POST'])
@auth.require_auth
def parkpow_recognize():
    """Recognize license plate using ParkPow API"""
    try:
        data = request.get_json()
        
        if not data or 'image' not in data:
            return jsonify({
                'success': False,
                'error': 'No image provided',
                'error_ar': 'لم يتم تقديم صورة'
            }), 400
        
        image_data = data.get('image')
        camera_id = data.get('camera_id')
        
        # Recognize plate
        result = parkpow_integration.recognize_plate(image_data, camera_id)
        
        # If recognition was successful, find vehicle and log
        if result.get('success') and result.get('results'):
            for plate_data in result['results']:
                plate_number = plate_data.get('plate') or plate_data.get('plate_number')
                
                if plate_number:
                    # Find vehicle in database
                    vehicle = parkpow_integration.find_vehicle_by_plate(plate_number)
                    
                    # Add vehicle info to result
                    plate_data['vehicle_info'] = vehicle
                    
                    # Log event
                    parkpow_integration.log_parkpow_event(
                        request.user['id'],
                        'recognition',
                        plate_number,
                        f"Confidence: {plate_data.get('confidence', 0)}"
                    )
        
        # Log audit
        database.log_audit(
            request.user['id'],
            'ParkPow plate recognition performed',
            ip_address=request.remote_addr
        )
        
        return jsonify(result)
    
    except Exception as e:
        app.logger.error(f'ParkPow recognition error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Failed to recognize plate',
            'error_ar': 'خطأ في تمييز اللوحة'
        }), 500


@app.route('/api/parkpow/record-violation', methods=['POST'])
@auth.require_auth
@auth.require_write_permission
def parkpow_record_violation():
    """Record a traffic violation"""
    try:
        data = request.get_json()
        
        if not data or 'plate_number' not in data:
            return jsonify({
                'success': False,
                'error': 'Plate number is required',
                'error_ar': 'رقم اللوحة مطلوب'
            }), 400
        
        plate_number = data.get('plate_number')
        violation_data = {
            'violation_type': data.get('violation_type', 'غير محدد'),
            'violation_date': data.get('violation_date', datetime.now().isoformat()),
            'location': data.get('location', ''),
            'description': data.get('description', ''),
            'fine_amount': data.get('fine_amount', 0)
        }
        
        # Record violation
        result = parkpow_integration.record_violation(
            plate_number,
            violation_data,
            request.user['id']
        )
        
        # Log audit
        database.log_audit(
            request.user['id'],
            f'ParkPow violation recorded for plate: {plate_number}',
            ip_address=request.remote_addr
        )
        
        return jsonify(result)
    
    except Exception as e:
        app.logger.error(f'ParkPow record violation error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Failed to record violation',
            'error_ar': 'خطأ في تسجيل المخالفة'
        }), 500


@app.route('/api/parkpow/repeat-offenders', methods=['GET'])
@auth.require_auth
def parkpow_repeat_offenders():
    """Get list of repeat offenders"""
    try:
        min_violations = request.args.get('min_violations', 3, type=int)
        
        offenders = parkpow_integration.get_repeat_offenders(min_violations)
        
        return jsonify({
            'success': True,
            'offenders': offenders,
            'count': len(offenders),
            'min_violations': min_violations
        })
    
    except Exception as e:
        app.logger.error(f'ParkPow repeat offenders error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Failed to get repeat offenders list',
            'error_ar': 'خطأ في جلب قائمة المخالفين المتكررين'
        }), 500


@app.route('/api/parkpow/webhook', methods=['POST'])
def parkpow_webhook():
    """Receive webhook data from ParkPow"""
    try:
        # Verify webhook token if configured
        token = request.headers.get('Authorization')
        expected_token = os.environ.get('PARKPOW_WEBHOOK_TOKEN', '')
        
        if expected_token and token != f'Token {expected_token}':
            return jsonify({
                'success': False,
                'error': 'Invalid webhook token',
                'error_ar': 'رمز webhook غير صالح'
            }), 401
        
        webhook_data = request.get_json()
        
        if not webhook_data:
            return jsonify({
                'success': False,
                'error': 'No data received',
                'error_ar': 'لم يتم استقبال بيانات'
            }), 400
        
        # Process webhook data
        result = parkpow_integration.process_webhook_data(webhook_data)
        
        return jsonify(result)
    
    except Exception as e:
        app.logger.error(f'ParkPow webhook error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Failed to process webhook',
            'error_ar': 'خطأ في معالجة webhook'
        }), 500


# ==================== Static File Serving ====================

@app.route('/')
def index():
    """Serve index page"""
    return send_from_directory('.', 'index.html')

@app.route('/<path:filename>')
def serve_static(filename):
    """Serve static files with security checks"""
    if not is_safe_path(filename):
        abort(403)
    
    safe_path = safe_join(BASE_DIR, filename)
    if safe_path is None or not os.path.isfile(safe_path):
        abort(404)
    
    return send_from_directory('.', filename)

# ==================== System Statistics ====================

@app.route('/api/system/stats')
def system_stats():
    """Get system statistics for validation report"""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Get counts for each table
        stats = {}
        
        # Buildings
        cursor.execute('SELECT COUNT(*) FROM buildings')
        stats['buildings'] = cursor.fetchone()[0]
        
        # Residents
        cursor.execute('SELECT COUNT(*) FROM residents')
        stats['residents'] = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM residents WHERE is_active = 1')
        stats['activeResidents'] = cursor.fetchone()[0]
        
        # Vehicles
        cursor.execute('SELECT COUNT(*) FROM vehicles')
        stats['vehicles'] = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM vehicles WHERE is_active = 1')
        stats['activeVehicles'] = cursor.fetchone()[0]
        
        # Traffic violations
        cursor.execute('SELECT COUNT(*) FROM traffic_violations')
        stats['violations'] = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM traffic_violations WHERE status = 'مفتوحة' OR status = 'open'")
        stats['openViolations'] = cursor.fetchone()[0]
        
        # Security incidents
        cursor.execute('SELECT COUNT(*) FROM security_incidents')
        stats['securityIncidents'] = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM security_incidents WHERE status = 'مفتوحة' OR status = 'open'")
        stats['openIncidents'] = cursor.fetchone()[0]
        
        # Visitors
        cursor.execute('SELECT COUNT(*) FROM visitors')
        stats['visitors'] = cursor.fetchone()[0]
        
        # Complaints
        cursor.execute('SELECT COUNT(*) FROM complaints')
        stats['complaints'] = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM complaints WHERE status = 'مفتوحة' OR status = 'open'")
        stats['openComplaints'] = cursor.fetchone()[0]
        
        # Users
        cursor.execute('SELECT COUNT(*) FROM users')
        stats['users'] = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM users WHERE is_active = 1')
        stats['activeUsers'] = cursor.fetchone()[0]
        
        # Active sessions
        cursor.execute('SELECT COUNT(*) FROM sessions WHERE expires_at > datetime("now")')
        stats['activeSessions'] = cursor.fetchone()[0]
        
        conn.close()
        
        return jsonify(stats)
        
    except Exception as e:
        app.logger.error(f'System stats error: {str(e)}')
        return jsonify({
            'error': 'Failed to get system statistics',
            'error_ar': 'فشل في الحصول على إحصائيات النظام'
        }), 500

# ==================== Health Check ====================

@app.route('/api/health')
def health_check():
    """
    Health check endpoint - Verifies database connectivity
    نقطة فحص الصحة - تتحقق من الاتصال بقاعدة البيانات
    """
    try:
        # Test database connection
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT 1')
        cursor.fetchone()
        conn.close()
        
        return jsonify({
            'status': 'healthy',
            'timestamp': datetime.now().isoformat(),
            'database': 'connected'
        })
    except Exception as e:
        # Log the error for debugging but don't expose details to users
        app.logger.error(f'Health check failed: {str(e)}')
        return jsonify({
            'status': 'unhealthy',
            'timestamp': datetime.now().isoformat(),
            'database': 'disconnected'
        }), 503

# ==================== Unified Dashboard API ====================

@app.route('/api/statistics')
def get_statistics():
    """Get statistics for unified dashboard"""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        stats = {}
        
        # Total residents
        cursor.execute('SELECT COUNT(*) FROM residents')
        stats['total_residents'] = cursor.fetchone()[0]
        
        # Total buildings
        cursor.execute('SELECT COUNT(*) FROM buildings')
        stats['total_buildings'] = cursor.fetchone()[0]
        
        # Total stickers
        cursor.execute('SELECT COUNT(*) FROM stickers')
        stats['total_stickers'] = cursor.fetchone()[0]
        
        # Total units (apartments)
        cursor.execute('SELECT COUNT(*) FROM apartments')
        stats['total_units'] = cursor.fetchone()[0]
        
        # Total parking spots
        cursor.execute('SELECT COUNT(*) FROM parking_spots')
        stats['total_parking'] = cursor.fetchone()[0]
        
        # Active violations
        cursor.execute("SELECT COUNT(*) FROM traffic_violations WHERE status = 'نشط' OR status = 'active'")
        stats['active_violations'] = cursor.fetchone()[0]
        
        conn.close()
        
        return jsonify({
            'success': True,
            'data': stats
        })
        
    except Exception as e:
        app.logger.error(f'Statistics error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Failed to get statistics',
            'error_ar': 'فشل في الحصول على الإحصائيات'
        }), 500

@app.route('/api/residents')
def get_residents():
    """Get all residents with their unit information"""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Get residents with building information
        query = """
        SELECT 
            r.id,
            r.name,
            r.national_id,
            r.email,
            r.phone,
            r.department,
            r.job_title,
            r.unit_number,
            r.move_in_date,
            r.move_out_date,
            r.is_active,
            b.building_number,
            b.name as building_name
        FROM residents r
        LEFT JOIN buildings b ON r.building_id = b.id
        ORDER BY r.id ASC
        """
        
        cursor.execute(query)
        rows = cursor.fetchall()
        
        # Format the data
        residents = []
        for row in rows:
            residents.append({
                'id': row[0],
                'name': row[1],
                'national_id': row[2],
                'email': row[3],
                'phone': row[4],
                'department': row[5],
                'job_title': row[6],
                'unit_number': row[7],
                'move_in_date': row[8],
                'move_out_date': row[9],
                'is_active': row[10],
                'building_number': row[11],
                'building_name': row[12]
            })
        
        conn.close()
        
        return jsonify({
            'success': True,
            'data': residents
        })
        
    except Exception as e:
        app.logger.error(f'Residents API error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Failed to get residents',
            'error_ar': 'فشل في الحصول على بيانات السكان'
        }), 500

@app.route('/api/violation-report')
def get_violation_report():
    """Get violation report with resident information"""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Get violations with resident information
        query = """
        SELECT 
            v.plate_number,
            COUNT(tv.id) as violation_count,
            v.vehicle_type,
            MAX(tv.violation_date) as latest_violation,
            r.name as resident_name,
            b.building_number,
            r.unit_number
        FROM traffic_violations tv
        LEFT JOIN vehicles v ON tv.vehicle_id = v.id
        LEFT JOIN residents r ON v.owner_id = r.id
        LEFT JOIN buildings b ON r.building_id = b.id
        WHERE v.plate_number IS NOT NULL
        GROUP BY v.plate_number
        ORDER BY violation_count DESC, latest_violation DESC
        """
        
        cursor.execute(query)
        rows = cursor.fetchall()
        
        # Format the data
        violations = []
        for row in rows:
            violations.append({
                'plateNumber': row[0],
                'violationCount': row[1],
                'vehicleType': row[2],
                'processingDate': row[3],
                'residentName': row[4],
                'buildingNumber': row[5],
                'unitNumber': row[6]
            })
        
        conn.close()
        
        return jsonify({
            'success': True,
            'data': violations
        })
        
    except Exception as e:
        app.logger.error(f'Violation report error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Failed to get violation report',
            'error_ar': 'فشل في الحصول على تقرير المخالفات'
        }), 500

# Note: This endpoint was renamed from '/api/residents' to '/api/residents-list' 
# to resolve a duplicate route conflict with another '/api/residents' endpoint at line ~843
# This is a pre-existing issue in the codebase that was preventing server startup
@app.route('/api/residents-list')
def get_residents_list():
    """Get all residents with building info"""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        building_id = request.args.get('building_id', type=int)
        is_active = request.args.get('is_active', type=int)
        
        query = '''
            SELECT r.id, r.name, r.national_id, r.email, r.phone, r.department, 
                   r.job_title, r.unit_number, r.move_in_date, r.move_out_date, 
                   r.is_active, b.name as building_name, b.building_number
            FROM residents r
            LEFT JOIN buildings b ON r.building_id = b.id
        '''
        
        conditions = []
        params = []
        
        if building_id:
            conditions.append('r.building_id = ?')
            params.append(building_id)
        
        if is_active is not None:
            conditions.append('r.is_active = ?')
            params.append(is_active)
        
        if conditions:
            query += ' WHERE ' + ' AND '.join(conditions)
        
        query += ' ORDER BY b.building_number, r.unit_number, r.name'
        
        cursor.execute(query, params)
        
        residents = []
        for row in cursor.fetchall():
            residents.append({
                'id': row[0],
                'name': row[1],
                'national_id': row[2],
                'email': row[3],
                'phone': row[4],
                'department': row[5],
                'job_title': row[6],
                'unit_number': row[7],
                'move_in_date': row[8],
                'move_out_date': row[9],
                'is_active': row[10],
                'building_name': row[11],
                'building_number': row[12]
            })
        
        conn.close()
        
        return jsonify({
            'success': True,
            'data': residents,
            'total': len(residents)
        })
        
    except Exception as e:
        app.logger.error(f'Residents API error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Failed to load residents data',
            'error_ar': 'فشل في تحميل بيانات السكان'
        }), 500

@app.route('/api/apartments')
def get_apartments():
    """Get all apartments with building info"""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        building_id = request.args.get('building_id', type=int)
        
        if building_id:
            cursor.execute('''
                SELECT a.id, a.unit_number, a.floor_number, a.unit_type, a.is_occupied,
                       b.name as building_name, b.building_number
                FROM apartments a
                JOIN buildings b ON a.building_id = b.id
                WHERE a.building_id = ?
                ORDER BY a.floor_number, a.unit_number
            ''', (building_id,))
        else:
            cursor.execute('''
                SELECT a.id, a.unit_number, a.floor_number, a.unit_type, a.is_occupied,
                       b.name as building_name, b.building_number
                FROM apartments a
                JOIN buildings b ON a.building_id = b.id
                ORDER BY b.building_number, a.floor_number, a.unit_number
            ''')
        
        apartments = []
        for row in cursor.fetchall():
            apartments.append({
                'id': row[0],
                'unit_number': row[1],
                'floor_number': row[2],
                'unit_type': row[3],
                'is_occupied': row[4],
                'building_name': row[5],
                'building_number': row[6]
            })
        
        conn.close()
        
        return jsonify({
            'success': True,
            'data': apartments,
            'total': len(apartments)
        })
        
    except Exception as e:
        app.logger.error(f'Apartments API error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Failed to load apartments data',
            'error_ar': 'فشل في تحميل بيانات الشقق'
        }), 500

@app.route('/api/parking-spots')
def get_parking_spots():
    """Get all parking spots with building and apartment info"""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        building_id = request.args.get('building_id', type=int)
        
        if building_id:
            cursor.execute('''
                SELECT p.id, p.spot_number, p.parking_area, p.is_occupied,
                       b.name as building_name, b.building_number,
                       a.unit_number
                FROM parking_spots p
                LEFT JOIN buildings b ON p.building_id = b.id
                LEFT JOIN apartments a ON p.apartment_id = a.id
                WHERE p.building_id = ?
                ORDER BY p.spot_number
            ''', (building_id,))
        else:
            cursor.execute('''
                SELECT p.id, p.spot_number, p.parking_area, p.is_occupied,
                       b.name as building_name, b.building_number,
                       a.unit_number
                FROM parking_spots p
                LEFT JOIN buildings b ON p.building_id = b.id
                LEFT JOIN apartments a ON p.apartment_id = a.id
                ORDER BY p.parking_area, p.spot_number
            ''')
        
        parking_spots = []
        for row in cursor.fetchall():
            parking_spots.append({
                'id': row[0],
                'spot_number': row[1],
                'parking_area': row[2],
                'is_occupied': row[3],
                'building_name': row[4],
                'building_number': row[5],
                'unit_number': row[6]
            })
        
        conn.close()
        
        return jsonify({
            'success': True,
            'data': parking_spots,
            'total': len(parking_spots)
        })
        
    except Exception as e:
        app.logger.error(f'Parking spots API error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Failed to load parking spots data',
            'error_ar': 'فشل في تحميل بيانات المواقف'
        }), 500

@app.route('/api/stickers')
def get_stickers():
    """Get all stickers data with resident information"""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT s.id, s.sticker_number, s.plate_number, s.vehicle_type, 
                   s.issue_date, s.expiry_date, s.status,
                   r.name as resident_name, r.national_id, r.phone,
                   r.department, r.job_title, r.unit_number,
                   b.name as building_name, b.building_number
            FROM stickers s
            LEFT JOIN residents r ON s.resident_id = r.id
            LEFT JOIN buildings b ON r.building_id = b.id
            ORDER BY s.sticker_number
        ''')
        
        stickers = []
        for row in cursor.fetchall():
            stickers.append({
                'id': row[0],
                'sticker_number': row[1],
                'plate_number': row[2],
                'vehicle_type': row[3],
                'issue_date': row[4],
                'expiry_date': row[5],
                'status': row[6],
                'resident_name': row[7],
                'national_id': row[8],
                'phone': row[9],
                'department': row[10],
                'job_title': row[11],
                'unit_number': row[12],
                'building_name': row[13],
                'building_number': row[14]
            })
        
        conn.close()
        
        return jsonify({
            'success': True,
            'data': stickers,
            'total': len(stickers)
        })
        
    except Exception as e:
        app.logger.error(f'Stickers API error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Failed to load stickers data',
            'error_ar': 'فشل في تحميل بيانات الملصقات'
        }), 500

@app.route('/api/buildings')
def get_buildings():
    """Get all buildings data"""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT id, name, building_number, address, total_units, total_floors, 
                   created_at, updated_at
            FROM buildings
            ORDER BY building_number
        ''')
        
        buildings = []
        for row in cursor.fetchall():
            buildings.append({
                'id': row[0],
                'name': row[1],
                'building_number': row[2],
                'address': row[3],
                'total_units': row[4],
                'total_floors': row[5],
                'created_at': row[6],
                'updated_at': row[7]
            })
        
        conn.close()
        
        return jsonify({
            'success': True,
            'data': buildings,
            'total': len(buildings)
        })
        
    except Exception as e:
        app.logger.error(f'Buildings API error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Failed to load buildings data',
            'error_ar': 'فشل في تحميل بيانات المباني'
        }), 500

@app.route('/api/comprehensive-reports')
def get_comprehensive_reports():
    """Get comprehensive system reports data"""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        reports = {}
        
        # Residents summary
        cursor.execute('SELECT COUNT(*) FROM residents WHERE is_active = 1')
        reports['totalResidents'] = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) FROM residents WHERE is_active = 0')
        reports['inactiveResidents'] = cursor.fetchone()[0]
        
        # Buildings summary
        cursor.execute('SELECT COUNT(*) FROM buildings')
        reports['totalBuildings'] = cursor.fetchone()[0]
        
        # Violations summary
        cursor.execute('SELECT COUNT(*) FROM traffic_violations')
        reports['totalViolations'] = cursor.fetchone()[0]
        
        cursor.execute("""
            SELECT COUNT(*) FROM traffic_violations 
            WHERE status IN ('pending', 'open', 'مفتوحة', 'معلقة')
        """)
        reports['openViolations'] = cursor.fetchone()[0]
        
        # Security incidents
        cursor.execute('SELECT COUNT(*) FROM security_incidents')
        reports['totalIncidents'] = cursor.fetchone()[0]
        
        cursor.execute("""
            SELECT COUNT(*) FROM security_incidents 
            WHERE status IN ('reported', 'open', 'مفتوحة')
        """)
        reports['openIncidents'] = cursor.fetchone()[0]
        
        # Complaints
        cursor.execute('SELECT COUNT(*) FROM complaints')
        reports['totalComplaints'] = cursor.fetchone()[0]
        
        cursor.execute("""
            SELECT COUNT(*) FROM complaints 
            WHERE status IN ('open', 'مفتوحة')
        """)
        reports['openComplaints'] = cursor.fetchone()[0]
        
        cursor.execute("""
            SELECT COUNT(*) FROM complaints 
            WHERE status IN ('resolved', 'محلولة', 'closed', 'مغلقة')
        """)
        reports['resolvedComplaints'] = cursor.fetchone()[0]
        
        # Vehicles and parking
        cursor.execute('SELECT COUNT(*) FROM vehicles WHERE is_active = 1')
        reports['activeVehicles'] = cursor.fetchone()[0]
        
        # Monthly occupancy trend (last 7 months)
        # TODO: Calculate from actual residents data when available
        reports['occupancyTrend'] = {
            'labels': ['يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو', 'يوليو'],
            'data': [85, 87, 89, 91, 88, 90, 92]  # Mock data - requires residents data
        }
        
        # Violations by type
        cursor.execute("""
            SELECT violation_type, COUNT(*) as count
            FROM traffic_violations
            GROUP BY violation_type
            ORDER BY count DESC
            LIMIT 5
        """)
        violation_types = cursor.fetchall()
        reports['violationsByType'] = {
            'labels': [row[0] for row in violation_types] if violation_types else ['وقوف ممنوع', 'عكس سير', 'مواقف ذوي الاحتياجات', 'عدم التقيد بالإشارات'],
            'data': [row[1] for row in violation_types] if violation_types else [15, 12, 6, 5]
        }
        
        # Security incidents trend (last 7 months)
        # TODO: Group by month when sufficient data exists
        reports['securityTrend'] = {
            'labels': ['يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو', 'يوليو'],
            'data': [8, 6, 10, 7, 9, 11, 12]  # Mock data - will be replaced with real aggregation
        }
        
        # Complaints trend
        # TODO: Calculate from actual complaints data with monthly grouping
        reports['complaintsTrend'] = {
            'labels': ['يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو', 'يوليو'],
            'new': [25, 30, 28, 35, 32, 29, 31],  # Mock data
            'resolved': [23, 28, 26, 33, 30, 27, 29]  # Mock data
        }
        
        # Residents by building
        cursor.execute("""
            SELECT b.name, COUNT(r.id) as count
            FROM buildings b
            LEFT JOIN residents r ON b.id = r.building_id AND r.is_active = 1
            GROUP BY b.id, b.name
            ORDER BY count DESC
            LIMIT 4
        """)
        residents_by_building = cursor.fetchall()
        reports['residentsByBuilding'] = {
            'labels': [row[0] for row in residents_by_building] if residents_by_building else ['المبنى 1', 'المبنى 2', 'المبنى 3', 'الفلل'],
            'data': [row[1] for row in residents_by_building] if residents_by_building else [65, 58, 72, 50]
        }
        
        conn.close()
        
        return jsonify({
            'success': True,
            'data': reports
        })
        
    except Exception as e:
        app.logger.error(f'Comprehensive reports error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Failed to load comprehensive reports data',
            'error_ar': 'فشل في تحميل بيانات التقارير الشاملة'
        }), 500

# ==================== Car Image Upload and Analysis Routes ====================

@app.route('/api/car-images/upload', methods=['POST'])
@auth.require_auth
def upload_car_images():
    """
    Upload multiple car images for analysis
    رفع عدة صور سيارات للتحليل
    """
    try:
        user = request.user
        
        # Check if files are present
        if 'images' not in request.files:
            return jsonify({
                'success': False,
                'error': 'No images provided',
                'error_ar': 'لم يتم تقديم أي صور'
            }), 400
        
        files = request.files.getlist('images')
        
        if not files or len(files) == 0:
            return jsonify({
                'success': False,
                'error': 'No images provided',
                'error_ar': 'لم يتم تقديم أي صور'
            }), 400
        
        uploaded_images = []
        analysis_results = []
        
        for file in files:
            if file and file.filename and allowed_image_file(file.filename):
                # Secure filename
                filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
                unique_filename = f"{timestamp}_{filename}"
                
                # Save image
                image_path = os.path.join(UPLOAD_FOLDER, unique_filename)
                file.save(image_path)
                
                # Create thumbnail
                thumbnail_filename = f"thumb_{unique_filename}"
                thumbnail_path = os.path.join(THUMBNAIL_FOLDER, thumbnail_filename)
                car_image_analyzer.create_thumbnail(image_path, thumbnail_path)
                
                # Save to database
                conn = database.get_db_connection()
                cursor = conn.cursor()
                
                cursor.execute('''
                    INSERT INTO car_images 
                    (original_filename, image_path, thumbnail_path, uploaded_by)
                    VALUES (?, ?, ?, ?)
                ''', (filename, image_path, thumbnail_path, user['id']))
                
                car_image_id = cursor.lastrowid
                conn.commit()
                conn.close()
                
                # Analyze image
                analysis = car_image_analyzer.analyze_car_image(image_path)
                
                if analysis['success']:
                    # Find matching vehicle
                    vehicle = None
                    vehicle_id = None
                    
                    if analysis.get('plate_number') and analysis['plate_number'] != 'غير محدد':
                        vehicle = car_image_analyzer.find_matching_vehicle(analysis['plate_number'])
                        if vehicle:
                            vehicle_id = vehicle['id']
                    
                    # Save analysis results
                    analysis_id = car_image_analyzer.save_car_analysis(
                        car_image_id,
                        analysis,
                        vehicle_id
                    )
                    
                    analysis_results.append({
                        'car_image_id': car_image_id,
                        'analysis_id': analysis_id,
                        'filename': filename,
                        'plate_number': analysis.get('plate_number'),
                        'vehicle_type': analysis.get('vehicle_type'),
                        'vehicle_color': analysis.get('vehicle_color'),
                        'confidence': analysis.get('plate_confidence'),
                        'matched_vehicle': vehicle is not None,
                        'vehicle_owner': vehicle.get('owner_name') if vehicle else None
                    })
                
                uploaded_images.append({
                    'id': car_image_id,
                    'filename': filename,
                    'path': image_path,
                    'thumbnail': thumbnail_path
                })
        
        # Log action
        database.log_audit(
            user['id'],
            f'Uploaded {len(uploaded_images)} car images for analysis',
            ip_address=request.remote_addr
        )
        
        return jsonify({
            'success': True,
            'message': f'Successfully uploaded and analyzed {len(uploaded_images)} images',
            'message_ar': f'تم رفع وتحليل {len(uploaded_images)} صورة بنجاح',
            'uploaded_images': uploaded_images,
            'analysis_results': analysis_results
        })
        
    except Exception as e:
        app.logger.error(f'Car image upload error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Failed to upload images. Please check the logs for details.',
            'error_ar': 'فشل في رفع الصور. يرجى التحقق من السجلات للحصول على التفاصيل.'
        }), 500


@app.route('/api/car-images/analysis', methods=['GET'])
@auth.require_auth
def get_car_analysis():
    """
    Get car analysis records
    الحصول على سجلات تحليل السيارات
    """
    try:
        records = car_data_exporter.get_car_analysis_records()
        
        return jsonify({
            'success': True,
            'data': records,
            'count': len(records)
        })
        
    except Exception as e:
        app.logger.error(f'Get car analysis error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Failed to get car analysis data',
            'error_ar': 'فشل في الحصول على بيانات تحليل السيارات'
        }), 500


@app.route('/api/car-images/export/<format>', methods=['POST'])
@auth.require_auth
def export_car_analysis(format):
    """
    Export car analysis data in specified format (excel, pdf, html)
    تصدير بيانات تحليل السيارات بالتنسيق المحدد
    """
    try:
        user = request.user
        
        # Get filter params from request
        filter_params = request.get_json() if request.is_json else None
        
        # Get data
        records = car_data_exporter.get_car_analysis_records(filter_params)
        
        if not records:
            return jsonify({
                'success': False,
                'error': 'No data to export',
                'error_ar': 'لا توجد بيانات للتصدير'
            }), 400
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if format == 'excel':
            filename = f'car_analysis_{timestamp}.xlsx'
            output_path = os.path.join(EXPORT_FOLDER, filename)
            
            success = car_data_exporter.export_to_excel(records, output_path)
            
            if success:
                database.log_audit(
                    user['id'],
                    f'Exported car analysis to Excel',
                    ip_address=request.remote_addr
                )
                
                return send_file(
                    output_path,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename
                )
        
        elif format == 'pdf':
            filename = f'car_analysis_{timestamp}.pdf'
            output_path = os.path.join(EXPORT_FOLDER, filename)
            
            success = car_data_exporter.export_to_pdf(records, output_path)
            
            if success:
                database.log_audit(
                    user['id'],
                    f'Exported car analysis to PDF',
                    ip_address=request.remote_addr
                )
                
                return send_file(
                    output_path,
                    mimetype='application/pdf',
                    as_attachment=True,
                    download_name=filename
                )
        
        elif format == 'html':
            filename = f'car_analysis_{timestamp}.html'
            output_path = os.path.join(EXPORT_FOLDER, filename)
            
            success = car_data_exporter.export_to_html(records, output_path)
            
            if success:
                database.log_audit(
                    user['id'],
                    f'Exported car analysis to HTML',
                    ip_address=request.remote_addr
                )
                
                return send_file(
                    output_path,
                    mimetype='text/html',
                    as_attachment=True,
                    download_name=filename
                )
        
        else:
            return jsonify({
                'success': False,
                'error': f'Unsupported export format: {format}',
                'error_ar': f'تنسيق التصدير غير مدعوم: {format}'
            }), 400
        
        return jsonify({
            'success': False,
            'error': 'Export failed',
            'error_ar': 'فشل التصدير'
        }), 500
        
    except Exception as e:
        app.logger.error(f'Export car analysis error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Failed to export data. Please check the logs for details.',
            'error_ar': 'فشل في تصدير البيانات. يرجى التحقق من السجلات للحصول على التفاصيل.'
        }), 500


@app.route('/api/car-images/thumbnail/<int:image_id>', methods=['GET'])
@auth.require_auth
def get_car_thumbnail(image_id):
    """
    Get thumbnail for a car image
    الحصول على صورة مصغرة لصورة السيارة
    """
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT thumbnail_path FROM car_images WHERE id = ?', (image_id,))
        row = cursor.fetchone()
        conn.close()
        
        if not row or not row[0]:
            return jsonify({
                'success': False,
                'error': 'Thumbnail not found',
                'error_ar': 'الصورة المصغرة غير موجودة'
            }), 404
        
        thumbnail_path = row[0]
        
        if not os.path.exists(thumbnail_path):
            return jsonify({
                'success': False,
                'error': 'Thumbnail file not found',
                'error_ar': 'ملف الصورة المصغرة غير موجود'
            }), 404
        
        return send_file(thumbnail_path, mimetype='image/jpeg')
        
    except Exception as e:
        app.logger.error(f'Get thumbnail error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Failed to get thumbnail',
            'error_ar': 'فشل في الحصول على الصورة المصغرة'
        }), 500


# ==================== Vehicle Report Export Routes ====================

@app.route('/api/vehicles/<int:vehicle_id>/export/<format>', methods=['GET'])
@auth.require_auth
def export_vehicle_report(vehicle_id, format):
    """
    Export single vehicle report with violations
    تصدير تقرير سيارة واحدة مع المخالفات
    """
    try:
        user = request.user
        
        # Get vehicle with violations
        vehicle = vehicle_report_exporter.get_vehicle_with_violations(vehicle_id)
        
        if not vehicle:
            return jsonify({
                'success': False,
                'error': 'Vehicle not found',
                'error_ar': 'السيارة غير موجودة'
            }), 404
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        plate = vehicle.get('plate_number', 'unknown').replace(' ', '_')
        
        if format == 'excel':
            filename = f'vehicle_report_{plate}_{timestamp}.xlsx'
            output_path = os.path.join(EXPORT_FOLDER, filename)
            
            success = vehicle_report_exporter.export_vehicle_to_excel(vehicle, output_path)
            
            if success:
                database.log_audit(
                    user['id'],
                    f'Exported vehicle report to Excel: {vehicle.get("plate_number")}',
                    ip_address=request.remote_addr
                )
                
                return send_file(
                    output_path,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename
                )
        
        elif format == 'pdf':
            filename = f'vehicle_report_{plate}_{timestamp}.pdf'
            output_path = os.path.join(EXPORT_FOLDER, filename)
            
            success = vehicle_report_exporter.export_vehicle_to_pdf(vehicle, output_path)
            
            if success:
                database.log_audit(
                    user['id'],
                    f'Exported vehicle report to PDF: {vehicle.get("plate_number")}',
                    ip_address=request.remote_addr
                )
                
                return send_file(
                    output_path,
                    mimetype='application/pdf',
                    as_attachment=True,
                    download_name=filename
                )
        
        elif format == 'html':
            filename = f'vehicle_report_{plate}_{timestamp}.html'
            output_path = os.path.join(EXPORT_FOLDER, filename)
            
            success = vehicle_report_exporter.export_vehicle_to_html(vehicle, output_path)
            
            if success:
                database.log_audit(
                    user['id'],
                    f'Exported vehicle report to HTML: {vehicle.get("plate_number")}',
                    ip_address=request.remote_addr
                )
                
                return send_file(
                    output_path,
                    mimetype='text/html',
                    as_attachment=True,
                    download_name=filename
                )
        
        else:
            return jsonify({
                'success': False,
                'error': f'Unsupported export format: {format}',
                'error_ar': f'تنسيق التصدير غير مدعوم: {format}'
            }), 400
        
        return jsonify({
            'success': False,
            'error': 'Export failed',
            'error_ar': 'فشل التصدير'
        }), 500
        
    except Exception as e:
        app.logger.error(f'Export vehicle report error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Server connection error. Please try again later.',
            'error_ar': 'خطأ في اتصال الخادم. يرجى المحاولة لاحقاً.'
        }), 500


@app.route('/api/vehicles/export-all/<format>', methods=['GET'])
@auth.require_auth
def export_all_vehicles_report(format):
    """
    Export all vehicles report
    تصدير تقرير جميع السيارات
    """
    try:
        user = request.user
        
        # Get all vehicles with violations
        vehicles = vehicle_report_exporter.get_all_vehicles_with_violations()
        
        if not vehicles:
            return jsonify({
                'success': False,
                'error': 'No vehicles found',
                'error_ar': 'لا توجد سيارات'
            }), 404
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if format == 'excel':
            filename = f'all_vehicles_report_{timestamp}.xlsx'
            output_path = os.path.join(EXPORT_FOLDER, filename)
            
            success = vehicle_report_exporter.export_all_vehicles_to_excel(vehicles, output_path)
            
            if success:
                database.log_audit(
                    user['id'],
                    f'Exported all vehicles report to Excel ({len(vehicles)} vehicles)',
                    ip_address=request.remote_addr
                )
                
                return send_file(
                    output_path,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename
                )
        
        elif format == 'html':
            filename = f'all_vehicles_report_{timestamp}.html'
            output_path = os.path.join(EXPORT_FOLDER, filename)
            
            success = vehicle_report_exporter.export_all_vehicles_to_html(vehicles, output_path)
            
            if success:
                database.log_audit(
                    user['id'],
                    f'Exported all vehicles report to HTML ({len(vehicles)} vehicles)',
                    ip_address=request.remote_addr
                )
                
                return send_file(
                    output_path,
                    mimetype='text/html',
                    as_attachment=True,
                    download_name=filename
                )
        
        else:
            return jsonify({
                'success': False,
                'error': f'Unsupported export format: {format}',
                'error_ar': f'تنسيق التصدير غير مدعوم: {format}'
            }), 400
        
        return jsonify({
            'success': False,
            'error': 'Export failed',
            'error_ar': 'فشل التصدير'
        }), 500
        
    except Exception as e:
        app.logger.error(f'Export all vehicles report error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Server connection error. Please try again later.',
            'error_ar': 'خطأ في اتصال الخادم. يرجى المحاولة لاحقاً.'
        }), 500


@app.route('/api/vehicles', methods=['GET'])
@auth.require_auth
def get_vehicles_list():
    """
    Get list of all vehicles with basic info
    الحصول على قائمة جميع السيارات
    """
    try:
        vehicles = vehicle_report_exporter.get_all_vehicles_with_violations()
        
        return jsonify({
            'success': True,
            'vehicles': vehicles,
            'total': len(vehicles)
        })
        
    except Exception as e:
        app.logger.error(f'Get vehicles list error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Server connection error. Please try again later.',
            'error_ar': 'خطأ في اتصال الخادم. يرجى المحاولة لاحقاً.'
        }), 500


# ==================== Historical Data Import Routes ====================

@app.route('/api/import/vehicles', methods=['POST'])
@auth.require_auth
def import_vehicles_data():
    """
    Import historical vehicles data from Excel
    استيراد البيانات التاريخية للسيارات من Excel
    """
    try:
        user = request.user
        
        # Check if user has admin role
        if user['role'] != 'admin':
            return jsonify({
                'success': False,
                'error': 'Unauthorized. Admin access required.',
                'error_ar': 'غير مصرح. يتطلب صلاحيات المدير.'
            }), 403
        
        # Check if file is provided
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'error': 'No file provided',
                'error_ar': 'لم يتم تقديم ملف'
            }), 400
        
        file = request.files['file']
        
        if not file.filename:
            return jsonify({
                'success': False,
                'error': 'No file selected',
                'error_ar': 'لم يتم اختيار ملف'
            }), 400
        
        # Check file extension
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({
                'success': False,
                'error': 'Invalid file format. Please upload Excel file (.xlsx or .xls)',
                'error_ar': 'تنسيق الملف غير صالح. يرجى رفع ملف Excel (.xlsx أو .xls)'
            }), 400
        
        # Save file temporarily
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'vehicles_import_{timestamp}.xlsx'
        file_path = os.path.join(EXPORT_FOLDER, filename)
        file.save(file_path)
        
        # Get merge mode from request
        merge_mode = request.form.get('merge_mode', 'true').lower() == 'true'
        
        # Import data
        stats = import_historical_vehicles.import_vehicles_from_excel(file_path, merge_mode)
        
        # Clean up temporary file
        try:
            os.remove(file_path)
        except:
            pass
        
        # Log audit
        database.log_audit(
            user['id'],
            f'Imported vehicles data: {stats["imported"]} imported, {stats["skipped"]} skipped, {stats["errors"]} errors',
            ip_address=request.remote_addr
        )
        
        return jsonify({
            'success': True,
            'message': 'Import completed',
            'message_ar': 'اكتمل الاستيراد',
            'stats': stats
        })
        
    except Exception as e:
        app.logger.error(f'Import vehicles error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Failed to import data. Please check file format.',
            'error_ar': 'فشل استيراد البيانات. يرجى التحقق من تنسيق الملف.'
        }), 500


@app.route('/api/import/violations', methods=['POST'])
@auth.require_auth
def import_violations_data():
    """
    Import historical violations data from Excel
    استيراد البيانات التاريخية للمخالفات من Excel
    """
    try:
        user = request.user
        
        # Check if user has admin role
        if user['role'] != 'admin':
            return jsonify({
                'success': False,
                'error': 'Unauthorized. Admin access required.',
                'error_ar': 'غير مصرح. يتطلب صلاحيات المدير.'
            }), 403
        
        # Check if file is provided
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'error': 'No file provided',
                'error_ar': 'لم يتم تقديم ملف'
            }), 400
        
        file = request.files['file']
        
        if not file.filename:
            return jsonify({
                'success': False,
                'error': 'No file selected',
                'error_ar': 'لم يتم اختيار ملف'
            }), 400
        
        # Check file extension
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({
                'success': False,
                'error': 'Invalid file format. Please upload Excel file (.xlsx or .xls)',
                'error_ar': 'تنسيق الملف غير صالح. يرجى رفع ملف Excel (.xlsx أو .xls)'
            }), 400
        
        # Save file temporarily
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'violations_import_{timestamp}.xlsx'
        file_path = os.path.join(EXPORT_FOLDER, filename)
        file.save(file_path)
        
        # Import data
        stats = import_historical_vehicles.import_violations_from_excel(file_path)
        
        # Clean up temporary file
        try:
            os.remove(file_path)
        except:
            pass
        
        # Log audit
        database.log_audit(
            user['id'],
            f'Imported violations data: {stats["imported"]} imported, {stats["skipped"]} skipped, {stats["errors"]} errors',
            ip_address=request.remote_addr
        )
        
        return jsonify({
            'success': True,
            'message': 'Import completed',
            'message_ar': 'اكتمل الاستيراد',
            'stats': stats
        })
        
    except Exception as e:
        app.logger.error(f'Import violations error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Failed to import data. Please check file format.',
            'error_ar': 'فشل استيراد البيانات. يرجى التحقق من تنسيق الملف.'
        }), 500


# ==================== Error Handlers ====================

@app.errorhandler(404)
def not_found(error):
    """Handle 404 errors"""
    return jsonify({
        'success': False,
        'error': 'Resource not found',
        'error_ar': 'المورد المطلوب غير موجود',
        'status': 404
    }), 404

@app.errorhandler(403)
def forbidden(error):
    """Handle 403 errors"""
    return jsonify({
        'success': False,
        'error': 'Access forbidden',
        'error_ar': 'الوصول غير مسموح به',
        'status': 403
    }), 403

@app.errorhandler(500)
def internal_error(error):
    """Handle 500 errors - Enhanced error message"""
    app.logger.error(f'Internal server error: {str(error)}')
    return jsonify({
        'success': False,
        'error': 'Internal server error. Please contact system administrator.',
        'error_ar': 'خطأ داخلي في الخادم. يرجى الاتصال بمسؤول النظام.',
        'status': 500
    }), 500

@app.errorhandler(Exception)
def handle_exception(error):
    """Handle all unhandled exceptions with professional error message"""
    app.logger.error(f'Unhandled exception: {str(error)}', exc_info=True)
    
    # Check if it's a connection error
    error_msg = str(error).lower()
    if 'connection' in error_msg or 'timeout' in error_msg or 'network' in error_msg:
        return jsonify({
            'success': False,
            'error': 'Server connection error. Please check your internet connection and try again.',
            'error_ar': 'خطأ في الاتصال بالخادم. يرجى التحقق من اتصال الإنترنت والمحاولة مرة أخرى.',
            'status': 503
        }), 503
    
    return jsonify({
        'success': False,
        'error': 'An unexpected error occurred. Please try again later.',
        'error_ar': 'حدث خطأ غير متوقع. يرجى المحاولة لاحقاً.',
        'status': 500
    }), 500

# ==================== Housing Report API ====================

@app.route('/api/housing-report/data')
def get_housing_report_data():
    """Get housing report data"""
    try:
        report_data = housing_report_generator.get_report_data()
        if report_data:
            return jsonify({
                'success': True,
                'data': report_data
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Failed to get report data'
            }), 500
    except Exception as e:
        app.logger.error(f'Housing report data error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Failed to retrieve report data. Please try again later.'
        }), 500


@app.route('/api/housing-report/export/pdf')
def export_housing_report_pdf():
    """Export housing report as PDF"""
    try:
        # Get report data
        report_data = housing_report_generator.get_report_data()
        if not report_data:
            return jsonify({
                'success': False,
                'error': 'Failed to get report data'
            }), 500
        
        # Generate PDF
        pdf_buffer = housing_report_generator.generate_pdf_report(report_data)
        if not pdf_buffer:
            return jsonify({
                'success': False,
                'error': 'Failed to generate PDF'
            }), 500
        
        # Create filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'housing_report_{timestamp}.pdf'
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        app.logger.error(f'PDF export error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Failed to generate PDF report. Please try again later.'
        }), 500


@app.route('/api/housing-report/export/word')
def export_housing_report_word():
    """Export housing report as Word document"""
    try:
        # Get report data
        report_data = housing_report_generator.get_report_data()
        if not report_data:
            return jsonify({
                'success': False,
                'error': 'Failed to get report data'
            }), 500
        
        # Generate Word document
        docx_buffer = housing_report_generator.generate_word_report(report_data)
        if not docx_buffer:
            return jsonify({
                'success': False,
                'error': 'Failed to generate Word document'
            }), 500
        
        # Create filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'housing_report_{timestamp}.docx'
        
        return send_file(
            docx_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        app.logger.error(f'Word export error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Failed to generate Word report. Please try again later.'
        }), 500

# ==================== Startup ====================

if __name__ == '__main__':
    # Cleanup expired sessions on startup
    auth.cleanup_expired_sessions()
    
    # Get configuration from environment
    debug_mode = os.environ.get('FLASK_DEBUG', 'True').lower() == 'true'
    port = int(os.environ.get('PORT', 5000))
    host = os.environ.get('HOST', '0.0.0.0')
    
    if debug_mode:
        print("=" * 60)
        print("⚠️  WARNING: Debug mode is enabled")
        print("⚠️  تحذير: وضع التصحيح مفعّل")
        print("⚠️  DO NOT use debug mode in production!")
        print("⚠️  لا تستخدم وضع التصحيح في بيئة الإنتاج!")
        print("=" * 60)
    
    app.run(host=host, port=port, debug=debug_mode)
